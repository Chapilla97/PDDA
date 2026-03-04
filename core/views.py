import matplotlib
matplotlib.use('Agg') # Importante para servidores web
import matplotlib.pyplot as plt
from docxtpl import InlineImage
from docx.shared import Mm
import io
import math 

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.conf import settings
from docxtpl import DocxTemplate
from docx import Document
import openpyxl
import os
from .models import Proyecto, RegistroActividad, Equipo, JustificacionSelectividad
from .forms import ProyectoForm, CrearUsuarioForm
import locale
from datetime import datetime
from django.contrib.auth.models import User, Group
from jinja2 import Environment, Undefined
from django.db.models import Q
from django.template.loader import get_template
from xhtml2pdf import pisa

# Configuración Locale
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8') 
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain')
    except:
        pass

def extraer_tags_de_archivo(archivo_en_memoria):
    nombre_archivo = archivo_en_memoria.name.lower()
    datos_encontrados = {}
    ERRORES_EXCEL = ['#DIV/0!', '#N/A', '#VALUE!', '#REF!', '#NAME?', '#NUM!', '#NULL!', 'nan']

    if nombre_archivo.endswith('.docx'):
        try:
            document = Document(archivo_en_memoria)
            namespace = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
            for element in document.element.body.iter():
                if element.tag.endswith('sdt'):
                    sdt_pr = element.find(namespace + 'sdtPr')
                    if sdt_pr is not None:
                        tag_element = sdt_pr.find(namespace + 'tag')
                        if tag_element is not None:
                            tag_name = tag_element.get(namespace + 'val')
                            sdt_content = element.find(namespace + 'sdtContent')
                            texto = "".join([t.text for t in sdt_content.iter(namespace + 't') if t.text]) if sdt_content else ""
                            datos_encontrados[tag_name] = texto
        except Exception: pass

    elif nombre_archivo.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(archivo_en_memoria, data_only=True)
            hojas_reales = wb.sheetnames
            for nombre_rango, objeto_definicion in wb.defined_names.items():
                if nombre_rango.startswith('_xlnm') or nombre_rango.startswith('Print_Area'): continue
                try:
                    try: destinations = list(objeto_definicion.destinations)
                    except: continue
                    for sheet_title, coord in destinations:
                        sheet_title_clean = sheet_title.strip("'")
                        if sheet_title_clean not in hojas_reales:
                            if len(hojas_reales) == 1: ws = wb[hojas_reales[0]]
                            else: continue
                        else: ws = wb[sheet_title_clean]
                        celda_limpia = coord.replace('$', '').split(':')[0]
                        try:
                            cell = ws[celda_limpia]
                            valor = cell.value
                            fmt = cell.number_format
                            
                            val_str_check = str(valor).strip().upper() if valor is not None else ""
                            if any(error in val_str_check for error in ERRORES_EXCEL):
                                valor = "" 
                            elif isinstance(valor, (int, float)):
                                if fmt and '%' in fmt:
                                    valor = f"{valor * 100:.2f}%"
                                elif fmt and ('0.' in fmt or '.' in fmt):
                                    try:
                                        if '.' in fmt:
                                            parte_decimal = fmt.split('.')[1]
                                            decimales = 0
                                            for char in parte_decimal:
                                                if char in ['0', '#', '?']: decimales += 1
                                                else: break
                                            valor = f"{valor:.{decimales}f}"
                                        else: valor = str(valor)
                                    except: valor = str(valor)
                                else: valor = str(valor)
                            datos_encontrados[nombre_rango] = str(valor) if valor is not None else ""
                        except: pass 
                except: pass
        except Exception: pass
    return datos_encontrados

def detectar_tipo_doc(filename):
    fname_clean = os.path.splitext(filename.lower())[0]
    fname_full = filename.lower() 
    
    # El orden importa: Primero buscamos los Protocolos para que no se confundan con las hojas
    mapa = {
        'Protocolo_Val': ['protocolo val', 'validacion', 'validación'],
        'Protocolo_Perfiles': ['protocolo perfiles', 'protocolo estudio', 'protocolo para estudios', 'ds-001'],
        '01_LS': ['01 ls', '01_ls', 'sistema'], 
        '02_IF': ['02 if', '02_if'],
        '03_Estabilidad': ['03', 'estabilidad'],
        '04_LM': ['04 lm', '04_lm', '04-lm'],
        '05_R': ['05 r', '05_r', '05-r'],
        '06_S': ['06 s', '06_s', '06-s'],
        'Factor_Similitud': ['factor', 'f2', 'similitud', 'ds-008'],
        'Porcentaje_Disuelto': ['disuelto', '% disuelto', 'ds-005'] # Quitamos la palabra "perfiles" de aquí
    }
    
    for clave, keywords in mapa.items():
        for kw in keywords:
            if kw in fname_clean: return clave
            if 'Protocolo' in clave and kw in fname_full: return clave
            
    return None
def evaluar_criterio(valor_str, operador, limite):
    try:
        if not valor_str or valor_str == "": return "-"
        limpio = str(valor_str).replace('%', '').strip()
        valor = float(limpio)
        limite = float(limite)
        if operador == '>=': return "Cumple" if valor >= limite else "No Cumple"
        if operador == '<=': return "Cumple" if valor <= limite else "No Cumple"
        if operador == '<': return "Cumple" if valor < limite else "No Cumple"
        if operador == '>': return "Cumple" if valor > limite else "No Cumple"
        return "-"
    except Exception:
        return "-"

@user_passes_test(lambda u: u.is_superuser)
def eliminar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    if request.method == 'POST':
        titulo = proyecto.titulo
        proyecto.delete()
        messages.error(request, f"🗑️ Proyecto '{titulo}' eliminado.")
    return redirect('lista_proyectos')

@login_required
def lista_proyectos(request):
    en_proceso = Proyecto.objects.filter(estado='en_proceso').order_by('-fecha_creacion')
    terminados = Proyecto.objects.filter(estado='terminado').order_by('-fecha_creacion')
    busqueda = request.GET.get('q')
    if busqueda:
        terminados = terminados.filter(Q(titulo__icontains=busqueda) | Q(nombre_analista__icontains=busqueda))
    return render(request, 'core/dashboard.html', {'en_proceso': en_proceso, 'terminados': terminados, 'busqueda': busqueda})

def registrar_log(proyecto, usuario, accion, comentario=None):
    RegistroActividad.objects.create(proyecto=proyecto, usuario=usuario, accion=accion, comentario=comentario)

@login_required
def crear_proyecto(request):
    try:
        grupo_analistas = Group.objects.get(name='Analistas')
        analistas = grupo_analistas.user_set.filter(is_active=True).order_by('first_name')
    except Group.DoesNotExist:
        analistas = User.objects.filter(is_active=True).order_by('username')
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        analista_id = request.POST.get('analista_id')
        if titulo:
            analista_user = User.objects.get(pk=analista_id) if analista_id else None
            nombre_analista = f"{analista_user.first_name} {analista_user.last_name}" if analista_user else (analista_user.username if analista_user else "Sin asignar")
            proyecto = Proyecto.objects.create(titulo=titulo, nombre_analista=nombre_analista, creado_por=request.user)
            registrar_log(proyecto, request.user, "Creó el proyecto")
            messages.success(request, "✨ Proyecto creado exitosamente.")
            return redirect('detalle_proyecto', proyecto_id=proyecto.id)
    return render(request, 'core/crear_proyecto.html', {'analistas': analistas})

@login_required
def detalle_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    docs_validacion = [('01_LS', 'Hoja: 01 LS'), ('02_IF', 'Hoja: 02 IF'), ('03_Estabilidad', 'Hoja: 03 Estabilidad'), ('04_LM', 'Hoja: 04 LM'), ('05_R', 'Hoja: 05 R'), ('06_S', 'Hoja: 06 S'), ('Protocolo_Val', 'Protocolo Validación')]
    docs_estudio = [('Factor_Similitud', 'Hoja: Factor Similitud'), ('Porcentaje_Disuelto', 'Hoja: % Disuelto'), ('Protocolo_Perfiles', 'Protocolo Perfiles')]

    if request.method == 'POST':
        if 'carga_masiva' in request.POST and proyecto.estado != 'terminado':
            archivos_lote = request.FILES.getlist('archivos_lote') 
            archivos_procesados = []
            archivos_ignorados = []
            for archivo in archivos_lote:
                clave_detectada = detectar_tipo_doc(archivo.name)
                if clave_detectada:
                    datos_nuevos = extraer_tags_de_archivo(archivo)
                    datos_nuevos['metadata'] = {'fecha': timezone.now().isoformat(), 'archivo': archivo.name, 'usuario': request.user.username}
                    target_dict = None
                    if clave_detectada in [d[0] for d in docs_validacion]:
                        if not isinstance(proyecto.datos_validacion, dict): proyecto.datos_validacion = {}
                        target_dict = proyecto.datos_validacion
                    elif clave_detectada in [d[0] for d in docs_estudio]:
                        if not isinstance(proyecto.datos_estudio, dict): proyecto.datos_estudio = {}
                        target_dict = proyecto.datos_estudio
                    
                    if target_dict is not None:
                        if clave_detectada not in target_dict: target_dict[clave_detectada] = {}
                        if isinstance(target_dict[clave_detectada], dict): target_dict[clave_detectada].update(datos_nuevos)
                        else: target_dict[clave_detectada] = datos_nuevos
                        archivos_procesados.append(f"{clave_detectada} ({archivo.name})")
                else:
                    archivos_ignorados.append(archivo.name)

            manual_update = False
            if request.POST.get('estabilidad_eau_horas'):
                if '03_Estabilidad' not in proyecto.datos_validacion: proyecto.datos_validacion['03_Estabilidad'] = {}
                campos_eau = ['estabilidad_eau_horas', 'estabilidad_eau_p_inicial', 'estabilidad_eau_p_final', 'estabilidad_eau_diferencia']
                for c in campos_eau:
                    val = request.POST.get(c)
                    if val: proyecto.datos_validacion['03_Estabilidad'][c] = val
                manual_update = True

            if archivos_procesados or manual_update:
                if isinstance(proyecto.datos_validacion, dict): proyecto.datos_validacion = dict(proyecto.datos_validacion)
                if isinstance(proyecto.datos_estudio, dict): proyecto.datos_estudio = dict(proyecto.datos_estudio)
                # -------------------------------------------------------
                
                proyecto.save()  # <--- Esta línea ya la tenías
                msg = f"✅ Procesados {len(archivos_procesados)} archivos."
                msg = f"✅ Procesados {len(archivos_procesados)} archivos."
                if archivos_ignorados: msg += f" ⚠️ Ignorados: {', '.join(archivos_ignorados)}"
                registrar_log(proyecto, request.user, f"Carga masiva: {', '.join(archivos_procesados)}")
                messages.success(request, msg)
            elif archivos_ignorados: messages.warning(request, f"⚠️ No se identificó ningún archivo.")
            else:
                if not manual_update: messages.warning(request, "⚠️ No seleccionaste ningún archivo.")
            return redirect('detalle_proyecto', proyecto_id=proyecto.id)

        if 'guardar_datos_generales' in request.POST and proyecto.estado != 'terminado':
            datos = {
                'fecha_inicio': request.POST.get('fecha_inicio'), 'fecha_fin': request.POST.get('fecha_fin'),
                'fecha_emision': request.POST.get('fecha_emision'), 'tecnica': request.POST.get('tecnica'),
                'equipo_modelo': request.POST.get('equipo_modelo'), 'selectividad_opcion': request.POST.get('selectividad_opcion'),
                'selectividad_texto': request.POST.get('selectividad_texto'),
            }
            def fmt_fecha(s):
                try:
                    dt = datetime.strptime(s, '%Y-%m-%d')
                    mes = dt.strftime('%B').capitalize()
                    return f"{dt.day} de {mes} de {dt.year}"
                except: return s
            try:
                dt_i = datetime.strptime(datos['fecha_inicio'], '%Y-%m-%d')
                dt_f = datetime.strptime(datos['fecha_fin'], '%Y-%m-%d')
                mes_f = dt_f.strftime('%B').capitalize()
                datos['periodo_validacion_txt'] = f"Del {dt_i.day} al {dt_f.day} de {mes_f} de {dt_f.year}"
            except: datos['periodo_validacion_txt'] = ""
            datos['fecha_emision_txt'] = fmt_fecha(datos['fecha_emision'])

            if datos['tecnica'] == 'croma':
                datos.update({'var_tecnica_nombre': "Cromatografía", 'var_tecnica_adj_pl': "Cromatográficas", 'var_tecnica_adj_sg': "Cromatográfico", 'var_unidades': "unidades de área", 'bloque_estabilidad_automuestreador': "Estabilidad en automuestreador"})
            elif datos['tecnica'] == 'espectro':
                datos.update({'var_tecnica_nombre': "Espectrofotometría", 'var_tecnica_adj_pl': "Espectrofotométricas", 'var_tecnica_adj_sg': "Espectrofotométrico", 'var_unidades': "unidades de absorbancia", 'bloque_estabilidad_automuestreador': ""})

            if not isinstance(proyecto.datos_validacion, dict): proyecto.datos_validacion = {}
            proyecto.datos_validacion['datos_generales'] = datos
            proyecto.save()
            registrar_log(proyecto, request.user, "Actualizó Configuración del Informe")
            messages.success(request, "💾 Configuración guardada.")
            return redirect('detalle_proyecto', proyecto_id=proyecto.id)

        if 'generar_informe_estudio' in request.POST:
            # === AQUÍ LLAMAMOS A LA NUEVA FUNCIÓN ===
            return generar_documento_estudio(proyecto, request)

        if 'subir_pdf_firmado' in request.POST:
            archivo = request.FILES.get('archivo_pdf')
            if archivo:
                proyecto.informe_final_firmado = archivo
                proyecto.fecha_subida_firmado = timezone.now()
                proyecto.save()
                registrar_log(proyecto, request.user, "Subió PDF Firmado Final")
                messages.success(request, "🚀 Proyecto Completado.")
                return redirect('detalle_proyecto', proyecto_id=proyecto.id)

        if 'enviar_comentario' in request.POST:
            texto = request.POST.get('comentario_texto')
            if texto:
                registrar_log(proyecto, request.user, "Comentario", texto)
                messages.info(request, "💬 Comentario agregado.")
                return redirect('detalle_proyecto', proyecto_id=proyecto.id)

        if 'finalizar_proyecto' in request.POST:
            proyecto.estado = 'terminado'
            proyecto.save()
            registrar_log(proyecto, request.user, "Finalizó el proyecto")
            return redirect('lista_proyectos')
        
        if 'reactivar_proyecto' in request.POST:
            proyecto.estado = 'en_proceso'
            proyecto.save()
            registrar_log(proyecto, request.user, "Reactivó el proyecto")
            return redirect('detalle_proyecto', proyecto_id=proyecto.id)

    historial = proyecto.actividades.all().order_by('-fecha')
    equipos = Equipo.objects.filter(activo=True)
    justificaciones = JustificacionSelectividad.objects.filter(activo=True)
    return render(request, 'core/detalle_proyecto.html', {'proyecto': proyecto, 'docs_validacion': docs_validacion, 'docs_estudio': docs_estudio, 'historial': historial, 'equipos': equipos, 'justificaciones': justificaciones})

class SilentUndefined(Undefined):
    def __str__(self): return ""

def generar_documento_descarga(contexto_datos, titulo_proyecto=""):
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'Plantilla_Informe.docx')
    if not os.path.exists(ruta_plantilla): return HttpResponse("Falta Plantilla_Informe.docx", status=404)
    doc = DocxTemplate(ruta_plantilla)
    contexto = {k: str(v) for k, v in contexto_datos.items()}
    jinja_env = Environment(undefined=SilentUndefined)
    doc.render(contexto, jinja_env)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="Informe_Validacion_Estudio_{titulo_proyecto}.docx"'
    doc.save(response)
    return response

# =========================================================================
# NUEVA FUNCIÓN PARA EL INFORME DE ESTUDIO (PERFILES)
# =========================================================================
def generar_documento_estudio(proyecto, request):
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'Plantilla_Estudio.docx')
    if not os.path.exists(ruta_plantilla): 
        return HttpResponse("Falta Plantilla_Estudio.docx en la raíz del proyecto", status=404)
        
    doc = DocxTemplate(ruta_plantilla)
    # jinja_env = Environment(undefined=SilentUndefined)
    
    # 1. Aplanar todos los datos de estudio
    datos = {}
    if isinstance(proyecto.datos_estudio, dict):
        for doc_datos in proyecto.datos_estudio.values():
            if isinstance(doc_datos, dict): datos.update(doc_datos)
            
    # Traemos también los generales por si los ocupas (fechas, etc)
    if isinstance(proyecto.datos_validacion, dict) and 'datos_generales' in proyecto.datos_validacion:
        datos.update(proyecto.datos_validacion['datos_generales'])
        
    # 2. Lógica Normativa NOM-177 (%CV)
    # IMPORTANTE: Asegúrate de nombrar los rangos en Excel como "cv_10_pba", "cv_20_pba", etc.
    tiempos = [10, 20, 30, 60, 90, 120]
    try:
        cv_10 = float(datos.get('cv_10_pba', 0))
        cv_subsecuentes = [float(datos.get(f'cv_{t}_pba', 0)) for t in tiempos[1:]]
        
        cumple_norma_cv = True
        if cv_10 > 20.0 or any(cv > 10.0 for cv in cv_subsecuentes):
            cumple_norma_cv = False
            
        if cumple_norma_cv:
            datos['texto_evaluacion'] = "se realiza la prueba de factor de similitud f2, demostrándose que los perfiles de disolución son similares al encontrar que el factor f2 es mayor a 50."
        else:
            datos['texto_evaluacion'] = "se emplea el método de Distancia de Mahalanobis debido a que la variabilidad de los datos excede los límites establecidos para aplicar f2."
    except Exception:
        datos['texto_evaluacion'] = "se realiza la prueba de factor de similitud f2 (Nota: Verifique los rangos de CV% en el Excel)."

    # 3. Reconstrucción dinámica de las tablas de perfiles (12 vasos)
    # Permite inyectarlos usando {% tr for perfil in perfiles_prueba %}
    perfiles_prueba = []
    perfiles_referencia = []
    for vaso in range(1, 13):
        perfil_pba = {'vaso': vaso}
        perfil_ref = {'vaso': vaso}
        for t in tiempos:
            perfil_pba[f't{t}'] = datos.get(f'v{vaso}_{t}_pba', '')
            perfil_ref[f't{t}'] = datos.get(f'v{vaso}_{t}_ref', '')
        perfiles_prueba.append(perfil_pba)
        perfiles_referencia.append(perfil_ref)
        
    datos['perfiles_prueba'] = perfiles_prueba
    datos['perfiles_referencia'] = perfiles_referencia

    # 4. Generación de Gráfica Promedio al Vuelo
    try:
        plt.figure(figsize=(8, 5))
        promedios_ref = [float(datos.get(f'prom_{t}_ref', 0)) for t in tiempos]
        promedios_prueba = [float(datos.get(f'prom_{t}_pba', 0)) for t in tiempos]
        
        plt.plot(tiempos, promedios_ref, marker='o', label='Referencia', color='#005a87')
        plt.plot(tiempos, promedios_prueba, marker='s', label='Prueba', color='#ff9900', linestyle='--')
        
        plt.title('Perfil de disolución promedio')
        plt.xlabel('Tiempo (minutos)')
        plt.ylabel('Porcentaje disuelto (%)')
        plt.legend()
        plt.grid(True, linestyle=':', alpha=0.5)
        plt.tight_layout()
        
        mem_grafica = io.BytesIO()
        plt.savefig(mem_grafica, format='png', dpi=150)
        mem_grafica.seek(0)
        datos['grafica_promedio'] = InlineImage(doc, mem_grafica, width=Mm(150))
        plt.close()
    except Exception as e:
        pass # Si hay error en los datos, genera el word sin romper el sistema

    contexto = {k: v for k, v in datos.items() if type(v) != dict}
    
    doc.render(contexto)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="Borrador_Estudio_{proyecto.titulo}.docx"'
    doc.save(response)
    
    registrar_log(proyecto, request.user, "Generó Borrador de Estudio con Gráfica y NOM-177")
    return response


@login_required
def generar_informe_validacion(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'Plantilla_Validacion.docx')
    if not os.path.exists(ruta_plantilla): return HttpResponse("Error: No encuentro 'Plantilla_Validacion.docx'", status=404)
    
    datos = {}
    if proyecto.datos_validacion:
        for doc in proyecto.datos_validacion.values():
            if isinstance(doc, dict): datos.update(doc)
    if 'datos_generales' in proyecto.datos_validacion: datos.update(proyecto.datos_validacion['datos_generales'])

    # --- MOTOR DE VALIDACIÓN AUTOMÁTICA ---
    reglas = [
        {'tag': 'ls_correlacion__r', 'op': '>=', 'lim': 0.99, 'dest': 'ls_r_cumple'},
        {'tag': 'ls_ERR', 'op': '<=', 'lim': 2.0, 'dest': 'ls_err_cumple'},
        {'tag': 'ls_precision_CV', 'op': '<=', 'lim': 2.0, 'dest': 'ls_cv_cumple'},
        {'tag': 'lm_correlacion__r_ref', 'op': '>=', 'lim': 0.99, 'dest': 'lm_r_ref_cumple'},
        {'tag': 'lm_ERR_ref', 'op': '<=', 'lim': 3.0, 'dest': 'lm_err_ref_cumple'},
        {'tag': 'lm_correlacion__r_pba', 'op': '>=', 'lim': 0.99, 'dest': 'lm_r_pba_cumple'},
        {'tag': 'lm_ERR_pba', 'op': '<=', 'lim': 3.0, 'dest': 'lm_err_pba_cumple'},
        {'tag': 'lm_CV_ref', 'op': '<=', 'lim': 3.0, 'dest': 'lm_cv_ref_cumple'},
        {'tag': 'r_CV_ref', 'op': '<=', 'lim': 3.0, 'dest': 'r_cv_ref_cumple'},
        {'tag': 'lm_CV_pba', 'op': '<=', 'lim': 3.0, 'dest': 'lm_cv_pba_cumple'},
        {'tag': 'r_CV_pba', 'op': '<=', 'lim': 3.0, 'dest': 'r_cv_pba_cumple'},
        {'tag': 'estabilidad_ea_diferencia', 'op': '<=', 'lim': 3.0, 'dest': 'est_ea_cumple'},
        {'tag': 'estabilidad_eau_diferencia', 'op': '<=', 'lim': 3.0, 'dest': 'est_eau_cumple'},
        {'tag': 'estabilidad_er_diferencia', 'op': '<=', 'lim': 3.0, 'dest': 'est_er_cumple'},
        {'tag': 'estabilidad_ess_diferencia', 'op': '<=', 'lim': 3.0, 'dest': 'est_ess_cumple'},
    ]
    for regla in reglas:
        valor = datos.get(regla['tag'], '')
        datos[regla['dest']] = evaluar_criterio(valor, regla['op'], regla['lim'])

    for tipo in ['ref', 'pba']:
        cumple_global = "Cumple"
        for i in range(1, 7):
            val = datos.get(f"lm_{i}_desviacion_{tipo}", '0')
            try:
                if abs(float(val)) > 3.0: cumple_global = "No Cumple"
            except: pass
        datos[f'exactitud_{tipo}_cumple'] = cumple_global

    for tipo in ['referencia', 'prueba']:
        cv = datos.get(f's_CV_{tipo}', '0')
        prom = datos.get(f's_promedio_{tipo}', '0')
        veredicto = "Cumple"
        try:
            if float(cv) > 3.0: veredicto = "No Cumple"
            if not (97.0 <= float(prom) <= 103.0): veredicto = "No Cumple"
        except: veredicto = "-"
        datos[f'selectividad_{tipo}_cumple'] = veredicto

    # --- INFLUENCIA DEL FILTRO DINÁMICA ---
    for tipo, prefijo in [('acrodisco', 'a'), ('canula', 'c')]:
        cumple_filtro = "Cumple"
        filtros_validos = 0
        texto_resultados = []
        for i in range(1, 6):
            nombre_filtro = datos.get(f'if_{tipo}_{i}', '').strip()
            dif_str = datos.get(f'if_diferencia_{prefijo}_{i}', '').strip()
            if nombre_filtro and dif_str and dif_str not in ['NA', '-', '']:
                filtros_validos += 1
                texto_resultados.append(f"{nombre_filtro} = {dif_str}")
                try:
                    if abs(float(dif_str)) > 2.0: cumple_filtro = "No Cumple"
                except: pass
        if filtros_validos == 0:
            cumple_filtro = "N/A"
            texto_final = "No se utilizaron filtros de este tipo."
        else:
            texto_final = "\n".join(texto_resultados)
        datos[f'if_{tipo}_cumple'] = cumple_filtro
        datos[f'if_{tipo}_texto_limpio'] = texto_final

    contexto = {k: str(v) for k, v in datos.items()}

    try:
        doc = DocxTemplate(ruta_plantilla)
        jinja_env = Environment(undefined=SilentUndefined)

        def crear_grafica_linealidad(prefijo_conc, prefijo_resp, prefijo_resul):
            try:
                plt.figure(figsize=(6.5, 4.5))
                colores = ['#005a87', '#00a3e0', '#63c0f5'] 
                todos_los_y = []
                for c in range(1, 4):
                    x_vals = []
                    y_vals = []
                    for n in range(1, 7): 
                        if prefijo_resp == "lm": key_resp = f"{prefijo_resp}_c{c}_n{n}_respuesta_{prefijo_resul}" 
                        else: key_resp = f"{prefijo_resp}_c{c}_n{n}_{prefijo_resul}" 
                        if "ls" in prefijo_conc: key_conc = f"{prefijo_conc}_1_{n}_concentracion"
                        else: key_conc = f"{prefijo_conc}_{n}_concentracion_{prefijo_resul}"
                        if key_resp in datos and key_conc in datos:
                            try:
                                val_x = float(datos[key_conc])
                                val_y = float(datos[key_resp])
                                x_vals.append(val_x)
                                y_vals.append(val_y)
                                todos_los_y.append(val_y)
                            except ValueError: continue
                    if x_vals and y_vals: plt.scatter(x_vals, y_vals, color=colores[c-1], marker='D', s=40, alpha=0.7, edgecolors='black', linewidths=0.5)
                try:
                    if "ls" in prefijo_conc:
                        pendiente = float(datos.get('ls_pendiente', 0))
                        intercepto = float(datos.get('ls_intercepto', 0))
                    else:
                        pendiente = float(datos.get(f'{prefijo_conc}_pendiente_{prefijo_resul}', 0))
                        intercepto = float(datos.get(f'{prefijo_conc}_intercepto_{prefijo_resul}', 0))
                    concs_teoricas = []
                    for n in range(1, 7):
                        try: 
                            if "ls" in prefijo_conc: k = f"{prefijo_conc}_1_{n}_concentracion"
                            else: k = f"{prefijo_conc}_{n}_concentracion_{prefijo_resul}"
                            concs_teoricas.append(float(datos.get(k, 0)))
                        except: pass
                    if concs_teoricas:
                        x_line = [min(concs_teoricas), max(concs_teoricas)]
                        y_line = [(m * pendiente + intercepto) for m in x_line]
                        plt.plot(x_line, y_line, color='black', linewidth=1.5, linestyle='-')
                        signo = "+" if intercepto >= 0 else ""
                        texto_ec = f'y = {pendiente:.4f}x {signo} {intercepto:.4f}'
                        plt.text(min(concs_teoricas) + 1, max(y_line)*0.9, texto_ec, fontsize=11, color='#333')
                except Exception: pass
                if todos_los_y:
                    max_y_real = max(todos_los_y)
                    limite_superior = math.ceil(max_y_real / 100) * 100
                    plt.ylim(0, limite_superior)
                plt.xlabel('Concentración (µg/mL)', fontsize=9, fontweight='bold')
                plt.ylabel('Respuesta (UA)', fontsize=9, fontweight='bold')
                plt.grid(True, linestyle=':', alpha=0.5)
                plt.tight_layout()
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', dpi=150)
                buffer.seek(0)
                plt.close()
                return InlineImage(doc, buffer, width=Mm(130))
            except Exception as e: return None

        img_ls = crear_grafica_linealidad("ls", "ls", "respuesta")
        if img_ls: contexto['grafica_linealidad_farmaco'] = img_ls
        img_ref = crear_grafica_linealidad("lm", "lm", "ref")
        if img_ref: contexto['grafica_linealidad_referencia'] = img_ref
        img_pba = crear_grafica_linealidad("lm", "lm", "pba")
        if img_pba: contexto['grafica_linealidad_prueba'] = img_pba
        
        doc.render(contexto, jinja_env)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        filename = f"Borrador_Validacion_{proyecto.titulo}.docx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        doc.save(response)
        registrar_log(proyecto, request.user, "Generó Borrador Validación con Verificación Automática y 3 Gráficas")
        return response
    except Exception as e: return HttpResponse(f"Error generando Word: {e}", status=500)

@user_passes_test(lambda u: u.is_superuser)
def administrar_usuarios(request):
    usuarios = User.objects.all().order_by('-date_joined')
    return render(request, 'core/lista_usuarios.html', {'usuarios': usuarios})

@user_passes_test(lambda u: u.is_superuser)
def crear_usuario_nuevo(request):
    if request.method == 'POST':
        form = CrearUsuarioForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            if User.objects.filter(username=data['username']).exists():
                messages.error(request, "El nombre de usuario ya existe.")
            else:
                user = User.objects.create(username=data['username'], first_name=data['first_name'], last_name=data['last_name'])
                if data['rol'] == 'analista':
                    user.set_unusable_password()
                    group, _ = Group.objects.get_or_create(name='Analistas')
                    user.groups.add(group)
                    messages.success(request, f"👤 Analista {user.first_name} creado.")
                elif data['rol'] == 'calidad':
                    if not data['password']:
                        messages.error(request, "Contraseña obligatoria.")
                        user.delete()
                        return render(request, 'core/crear_usuario.html', {'form': form})
                    user.set_password(data['password'])
                    group, _ = Group.objects.get_or_create(name='Calidad')
                    user.groups.add(group)
                    messages.success(request, f"🛡️ Usuario Calidad {user.first_name} creado.")
                user.save()
                return redirect('administrar_usuarios')
    else: form = CrearUsuarioForm()
    return render(request, 'core/crear_usuario.html', {'form': form})

@login_required
def exportar_audit_trail_pdf(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    historial = proyecto.actividades.all().order_by('-fecha')
    context = {'proyecto': proyecto, 'historial': historial}
    template = get_template('core/pdf_audit_trail.html')
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    filename = f"AuditTrail_{proyecto.id}_{timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: return HttpResponse('Error PDF', status=500)
    return response